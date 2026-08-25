"""Load data/ParcelPilot_Assessment_Data.xlsx into accounts, orders, tickets."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.core.config import settings
from app.integrations.db.postgres import connection

WORKBOOK = ROOT / "data" / "ParcelPilot_Assessment_Data.xlsx"

TABLES = ("accounts", "orders", "tickets")


def _norm_header(name: str) -> str:
    return str(name or "").strip().lower().replace(" ", "_")


def _cell(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _sheet_rows(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_name:
                sheet_name = name
                break
        else:
            return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {_norm_header(headers[i]): _cell(row[i]) for i in range(len(headers)) if headers[i]}
        if record:
            out.append(record)
    return out


def _insert_rows(table: str, rows: list[dict]) -> int:
    if not rows:
        return 0
    cols = list(rows[0].keys())
    placeholders = ", ".join(f"%({c})s" for c in cols)
    col_list = ", ".join(cols)
    sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) ON CONFLICT DO NOTHING"
    with connection() as conn:
        with conn.cursor() as cur:
            for row in rows:
                cur.execute(sql, row)
        conn.commit()
    return len(rows)


def main() -> None:
    if not settings.database_url:
        print("DATABASE_URL is not set in .env")
        print()
        print("Supabase: Project Settings -> Database -> Connection string -> URI")
        print("Example shape:")
        print("  DATABASE_URL=postgresql://postgres.[ref]:[YOUR-PASSWORD]@aws-0-[region].pooler.supabase.com:5432/postgres")
        print()
        print("Or run generated SQL in the Supabase SQL editor:")
        print("  uv run python scripts/gen_seed_sql.py")
        print("  then paste scripts/_seed.sql")
        sys.exit(1)
    if not WORKBOOK.exists():
        print(f"Missing workbook: {WORKBOOK}")
        print("Drop ParcelPilot_Assessment_Data.xlsx into data/ then re-run.")
        sys.exit(1)
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    total = 0
    try:
        for table in TABLES:
            rows = _sheet_rows(wb, table)
            if not rows:
                print(f"  skip {table}: no sheet or empty")
                continue
            n = _insert_rows(table, rows)
            total += n
            print(f"  {table}: {n} rows")
    except Exception as exc:
        msg = str(exc)
        print(f"Database error: {msg}")
        if "getaddrinfo failed" in msg or "11001" in msg:
            print()
            print("Windows/IPv4: use the Session pooler URI from Supabase Connect (port 5432),")
            print("not the direct db.*.supabase.co host. URL-encode @ in passwords as %40.")
        elif "password authentication failed" in msg.lower():
            print()
            print("Wrong DB password. Supabase → Settings → Database → reset password,")
            print("then copy the Session pooler URI into DATABASE_URL.")
        print()
        print("Tip: data may already be in Supabase — check Table Editor or run verify_schema.py.")
        sys.exit(1)
    finally:
        wb.close()
    print(f"Done ({total} rows attempted).")


if __name__ == "__main__":
    main()
